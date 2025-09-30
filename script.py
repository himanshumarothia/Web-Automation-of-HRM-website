from selenium import webdriver
from selenium.webdriver.common.by import By 
from openpyxl import load_workbook
import time


driver = webdriver.Chrome()
url = "https://opensource-demo.orangehrmlive.com/web/index.php/auth/login"  
driver.get(url)
driver.maximize_window()
time.sleep(2)                #delay of 2sec


#locating username and password element
username = driver.find_element(By.XPATH,"//input[@placeholder='Username']")    
username.send_keys("Admin")                           #sending the value
time.sleep(1)
password = driver.find_element(By.XPATH,"//input[@placeholder='Password']")
password.send_keys("admin123")
time.sleep(1)


#login button 
login = driver.find_element(By.XPATH,"//button[normalize-space()='Login']")
login.click()
time.sleep(5)


#mouse hover to PIM
pim = driver.find_element(By.XPATH,"//span[normalize-space()='PIM']")
pim.click()
time.sleep(4)

#add button
add_button = driver.find_element(By.XPATH,"//button[normalize-space()='Add']")
add_button.click()
time.sleep(5)

#importing employee data from excel sheet and creating new employees
workbook=load_workbook('New Employees Data.xlsx')
sheet = workbook.active                                   #selecting active sheet

for row in sheet.iter_rows(min_row=2,max_row=sheet.max_row,values_only=True) :
    firstname = row[0]
    middlename = row[1]
    lastname = row[2]


    driver.find_element(By.XPATH,"//input[@placeholder='First Name']").send_keys(firstname)
    driver.find_element(By.XPATH,"//input[@placeholder='Middle Name']").send_keys(middlename)
    driver.find_element(By.XPATH,"//input[@placeholder='Last Name']").send_keys(lastname)
    time.sleep(1)
    driver.find_element(By.XPATH,"//button[normalize-space()='Save']").click()
    time.sleep(13)
    
    add_emp = driver.find_element(By.XPATH,"//a[normalize-space()='Add Employee']").click()
    time.sleep(6)



# Go to employee list and verify by searching
emp_list = driver.find_element(By.XPATH,"//a[normalize-space()='Employee List']").click()
time.sleep(5)

empfirstname = driver.find_element(By.XPATH,"(//input[@placeholder='Type for hints...'])[1]").send_keys('Shyam')
search_button = driver.find_element(By.XPATH,"//button[normalize-space()='Search']").click()

#wait for table to load and check 
time.sleep(4)
# searching for employee Ram 
results = driver.find_elements(By.XPATH,"//div[@class='oxd-table-body']//div[contains(text(), 'Shyam')]")   
if results:
    print("Name Verified")
else :
    print("Employee Not Found")    
        


#logout from dashboard
profile = driver.find_element(By.XPATH,"//p[@class='oxd-userdropdown-name']").click()
time.sleep(3)
logout_button = driver.find_element(By.XPATH,"//a[normalize-space()='Logout']").click()


#close the browser
driver.quit()
